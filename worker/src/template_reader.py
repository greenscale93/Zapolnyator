"""
Чтение данных из шаблона Excel-отчёта.

Содержит:
- get_template_offices — список офисов из шаблона
- find_row_by_name — поиск строки по тексту в колонке
- find_month_column — поиск колонки месяца в шапке
- read_cell_at — чтение значения из ячейки (с учётом объединённых)
"""
import logging
import os
import tempfile
import zipfile
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter
import msoffcrypto

logger = logging.getLogger(__name__)


def _open_workbook(template_path: str, data_only: bool = True):
    """Открывает workbook с поддержкой зашифрованных файлов."""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")

    wb = None
    try:
        wb = openpyxl.load_workbook(template_path, data_only=data_only)
    except zipfile.BadZipFile:
        try:
            with open(template_path, 'rb') as f:
                file = msoffcrypto.OfficeFile(f)
                if file.is_encrypted():
                    file.load_key(password="987456")
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        file.decrypt(tmp)
                        tmp_path = tmp.name
                    wb = openpyxl.load_workbook(tmp_path, data_only=data_only)
                    os.unlink(tmp_path)
                else:
                    raise
        except Exception:
            raise ValueError(
                "Не удалось прочитать шаблон (возможно, повреждён или неверный пароль)"
            )
    if wb is None:
        raise ValueError("Не удалось открыть файл шаблона")
    return wb


def get_sheet(ws_name: str = "Отчетность БИТ 2026"):
    """Декоратор, открывающий workbook и передающий ws в функцию."""
    # Эта функция используется внутри open_workbook_and_sheet
    pass


def open_workbook_and_sheet(template_path: str, sheet_name: str = "Отчетность БИТ 2026", data_only: bool = True):
    """Открывает workbook и лист, возвращает (wb, ws)."""
    wb = _open_workbook(template_path, data_only=data_only)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Лист '{sheet_name}' не найден в шаблоне")
    ws = wb[sheet_name]
    return wb, ws


def get_template_offices(template_path: str) -> list:
    """
    Читает шаблон Excel и возвращает отсортированный список названий
    офисов/подразделений из листа 'Отчетность БИТ 2026'.
    """
    wb = _open_workbook(template_path, data_only=True)
    if "Отчетность БИТ 2026" not in wb.sheetnames:
        wb.close()
        raise ValueError("Лист 'Отчетность БИТ 2026' не найден в шаблоне")

    ws = wb["Отчетность БИТ 2026"]
    offices = set()
    for row in range(16, 33):
        val = ws.cell(row=row, column=3).value
        if val and str(val).strip():
            offices.add(str(val).strip())
    for row in range(34, 36):
        val = ws.cell(row=row, column=3).value
        if val and str(val).strip():
            offices.add(str(val).strip())
    wb.close()
    return sorted(list(offices))


def find_row_by_name(ws, column: int, name: str) -> Optional[int]:
    """
    Находит номер строки, в которой в указанной колонке (1-based)
    содержится искомый текст.
    """
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column).value
        if cell_value and name.strip() in str(cell_value).strip():
            logger.info(f"Строка '{name}' найдена в {get_column_letter(column)}{row}")
            return row
    logger.warning(f"Строка '{name}' не найдена в колонке {get_column_letter(column)}")
    return None


def find_month_column(ws, month_name: str, search_row: int = 2) -> Optional[int]:
    """
    Находит номер колонки (1-based), в которой в указанной строке
    содержится название месяца.
    """
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=search_row, column=col).value
        if cell_value and month_name.strip() in str(cell_value).strip():
            logger.info(f"Месяц '{month_name}' найден в {get_column_letter(col)}{search_row}")
            return col
    logger.warning(f"Месяц '{month_name}' не найден в строке {search_row}")
    return None


def read_cell_at(ws, row: int, column: int):
    """
    Читает значение из ячейки. Если ячейка пустая, но входит
    в объединённый диапазон — возвращает значение из верхней левой
    ячейки объединения.
    """
    cell = ws.cell(row=row, column=column)
    if cell.value is not None:
        return cell.value

    # Проверяем, не входит ли ячейка в объединённый диапазон
    for merged_range in ws.merged_cells.ranges:
        if (cell.coordinate in merged_range
                or (merged_range.min_row <= row <= merged_range.max_row
                    and merged_range.min_col <= column <= merged_range.max_col)):
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left.value
    return None
