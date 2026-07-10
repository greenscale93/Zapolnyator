import logging
import os
import json
import re
import pandas as pd
from fastapi import FastAPI
from pydantic import BaseModel
from typing import Any, Dict, Optional
from src.mxl_parser import parse_mxl, convert_mxl_to_csv
from src.excel_processor import read_excel_structure, apply_sheet_mapping
import openpyxl
import tempfile
import zipfile
import msoffcrypto

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

class ToolRequest(BaseModel):
    tool: str
    arguments: Dict[str, Any]

class ToolResponse(BaseModel):
    status: str
    result: Optional[Dict[str, Any]] = None
    error_message: Optional[str] = None

def sanitize_data(data):
    if isinstance(data, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', data)
    elif isinstance(data, dict):
        return {sanitize_data(k): sanitize_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_data(item) for item in data]
    else:
        return data

@app.post("/api/v1/tool")
async def call_tool(request: ToolRequest):
    logger.info(f"Calling tool: {request.tool} with args: {request.arguments}")
    try:
        if request.tool == "parse_mxl":
            result = await parse_mxl(request.arguments["file_path"])
            if result.get("status") == "error":
                return ToolResponse(status="error", error_message=result.get("error_message"))
            return ToolResponse(status="success", result=result.get("result"))
        
        elif request.tool == "convert_mxl_to_csv":
            result = await convert_mxl_to_csv(request.arguments["file_path"])
            if result.get("status") == "error":
                return ToolResponse(status="error", error_message=result.get("error_message"))
            return ToolResponse(status="success", result=result.get("result"))
        
        elif request.tool == "read_excel_data":
            file_path = request.arguments["file_path"]
            sheet_name = request.arguments.get("sheet_name")
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
                df = df.where(pd.notnull(df), None)
                data = df.to_dict(orient='records')
                columns = df.columns.tolist()
                data = sanitize_data(data)
                return ToolResponse(status="success", result={
                    "data": data,
                    "columns": columns,
                    "rows": len(data)
                })
            except Exception as e:
                logger.exception("read_excel_data error")
                return ToolResponse(status="error", error_message=str(e))
        
        elif request.tool == "apply_sheet_mapping":
            source_path = request.arguments["source_path"]
            template_path = request.arguments["template_path"]
            sheet_name = request.arguments["sheet_name"]
            mapping = request.arguments["mapping"]
            month = request.arguments["month"]
            year = request.arguments["year"]
            password = request.arguments.get("password", "987456")
            output_path = request.arguments.get("output_path")
            result = await apply_sheet_mapping(source_path, template_path, sheet_name, mapping, month, year, password, output_path)
            if result.get("status") == "error":
                return ToolResponse(status="error", error_message=result.get("error_message"))
            return ToolResponse(status="success", result={"output_path": result["output_path"], "rows_added": result.get("rows_added")})
        
        elif request.tool == "read_excel_structure":
            file_path = request.arguments["file_path"]
            result = await read_excel_structure(file_path)
            if result.get("status") == "error":
                return ToolResponse(status="error", error_message=result.get("error_message"))
            return ToolResponse(status="success", result=result.get("result"))

        elif request.tool == "read_vz_empty_contractors":
            source_path = request.arguments["source_path"]
            if not os.path.exists(source_path):
                return ToolResponse(status="error", error_message=f"Source file not found: {source_path}")
            try:
                df = pd.read_excel(source_path, header=0)
                # Фильтр: только строки Взаиморасчет и направления Доход/Расход
                mask = (df["ТипЗаписи"].astype(str) == "Взаиморасчет") & (df["Направление"].astype(str).isin(["Доход", "Расход"]))
                df_vz = df[mask]
                # Пустые ПодразделениеКонтрагентДляОтчета
                empty_mask = df_vz["ПодразделениеКонтрагентДляОтчета"].isna() | (df_vz["ПодразделениеКонтрагентДляОтчета"].astype(str).str.strip() == "")
                contractors = df_vz.loc[empty_mask, "ПодразделениеКонтрагент"].dropna().unique().tolist()
                return ToolResponse(status="success", result={"contractors": contractors})
            except Exception as e:
                logger.exception("read_vz_empty_contractors error")
                return ToolResponse(status="error", error_message=str(e))
        
        elif request.tool == "read_template_offices":
            template_path = request.arguments["template_path"]
            if not os.path.exists(template_path):
                return ToolResponse(status="error", error_message=f"Template file not found: {template_path}")
            try:
                # Пытаемся открыть с расшифровкой, если файл зашифрован
                wb = None
                try:
                    # Сначала обычное открытие
                    wb = openpyxl.load_workbook(template_path, data_only=True)
                except zipfile.BadZipFile:
                    # Возможно файл зашифрован, пробуем расшифровать паролем "987456"
                    try:
                        with open(template_path, 'rb') as f:
                            file = msoffcrypto.OfficeFile(f)
                            if file.is_encrypted():
                                file.load_key(password="987456")
                                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                                    file.decrypt(tmp)
                                    tmp_path = tmp.name
                                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                                os.unlink(tmp_path)
                            else:
                                raise  # если не зашифрован, перехватим ниже
                    except:
                        raise ValueError("Не удалось прочитать шаблон (возможно, повреждён или неверный пароль)")

                if wb is None:
                    return ToolResponse(status="error", error_message="Не удалось открыть файл шаблона")

                if "Отчетность БИТ 2026" not in wb.sheetnames:
                    wb.close()
                    return ToolResponse(status="error", error_message="Лист 'Отчетность БИТ 2026' не найден в шаблоне")

                ws = wb["Отчетность БИТ 2026"]
                offices = set()
                for row in range(13, 17):
                    val = ws.cell(row=row, column=3).value
                    if val and str(val).strip():
                        offices.add(str(val).strip())
                for row in range(34, 36):
                    val = ws.cell(row=row, column=3).value
                    if val and str(val).strip():
                        offices.add(str(val).strip())
                wb.close()
                return ToolResponse(status="success", result={"offices": sorted(list(offices))})
            except Exception as e:
                logger.exception("read_template_offices error")
                return ToolResponse(status="error", error_message=str(e))
                    
        else:
            return ToolResponse(status="error", error_message=f"Unknown tool: {request.tool}")
    
    except Exception as e:
        logger.exception("Tool execution error")
        return ToolResponse(status="error", error_message=str(e))

@app.get("/health")
async def health():
    return {"status": "ok"}