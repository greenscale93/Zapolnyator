"""
Запись/чтение значений в шаблон отчёта по конфигурации из rules.json.

Поддерживает:
- write_values (тип ffot) — расчёт + запись в ячейку
- read_values — чтение значения из ячейки
"""
import logging
import re
import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Optional, List

from src.template_reader import (
    find_row_by_name, find_month_column, read_cell_at, open_workbook_and_sheet
)

logger = logging.getLogger(__name__)

MONTHS_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}


def get_month_name(month) -> str:
    """Возвращает название месяца по номеру (1-12) или строке."""
    if isinstance(month, str):
        return month
    return MONTHS_RU.get(month, str(month))


def get_month_number(month_name: str) -> int:
    """Конвертирует название месяца в число."""
    rev_map = {v: k for k, v in MONTHS_RU.items()}
    num = rev_map.get(month_name)
    if num is None:
        raise ValueError(f"Неизвестное название месяца: {month_name}")
    return num


def get_target_month_name(month, offset: int) -> str:
    """Возвращает название месяца со смещением (0 = текущий, -1 = предыдущий)."""
    if isinstance(month, str):
        month_num = get_month_number(month)
    else:
        month_num = month
    target = month_num + offset
    # Циклический сдвиг в пределах года
    while target < 1:
        target += 12
    while target > 12:
        target -= 12
    return MONTHS_RU[target]


def _clean_numeric(value):
    """Очищает значение от неразрывных пробелов и приводит к float."""
    if isinstance(value, str):
        cleaned = re.sub(r'[^\d,.-]', '', value.replace('\u00a0', '').replace(' ', ''))
        cleaned = cleaned.replace(',', '.')
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


# ===================== WRITE VALUES =====================


async def _write_ffot(
    source_path: str,
    template_path: str,
    config: dict,
    month,
    year: int
) -> dict:
    """
    Обрабатывает write_value типа 'ffot':
    - читает source, фильтрует ФактическийФОТ
    - суммирует Оклад+Премия с учётом исключений
    - находит целевую ячейку по config
    - записывает значение
    """
    # 1. Расчёт суммы
    exclude_emp = config.get("exclude_employee", {})
    ffot_value = _calculate_ffot_sum(source_path, exclude_employee=exclude_emp)

    # 2. Открываем шаблон (без data_only, чтобы не потерять формулы)
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception:
        import tempfile
        import msoffcrypto
        with open(template_path, 'rb') as f:
            file = msoffcrypto.OfficeFile(f)
            if file.is_encrypted():
                file.load_key(password="987456")
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    file.decrypt(tmp)
                    tmp_path = tmp.name
                wb = openpyxl.load_workbook(tmp_path)
                os.unlink(tmp_path)
            else:
                raise

    ws = wb["Отчетность БИТ 2026"]

    # 3. Целевая колонка
    target_month = get_target_month_name(month, config.get("month_offset", -1))
    col_idx = find_month_column(
        ws, target_month, config.get("month_row", 2)
    )
    if col_idx is None:
        wb.close()
        raise ValueError(f"Не найден месяц '{target_month}' в строке {config.get('month_row', 2)}")

    # 4. Целевая строка
    row_num = find_row_by_name(
        ws, config.get("row_column", 3), config.get("row_name", "")
    )
    if row_num is None:
        wb.close()
        raise ValueError(
            f"Не найдена строка '{config.get('row_name')}' "
            f"в колонке {get_column_letter(config.get('row_column', 3))}"
        )

    # 5. Запись
    cell_ref = f"{get_column_letter(col_idx)}{row_num}"
    ws.cell(row=row_num, column=col_idx).value = ffot_value

    from src.excel_processor import _save_and_fix_formats
    await _save_and_fix_formats(wb, template_path)

    logger.info(f"Записано {config.get('key')}: {ffot_value} в {cell_ref}")
    return {"value": ffot_value, "cell": cell_ref, "label": config.get("label", ""), "format": config.get("format", "amount")}


async def _write_admin_count(
    source_path: str,
    template_path: str,
    config: dict,
    month,
    year: int
) -> dict:
    """
    Обрабатывает write_value типа 'admin_count':
    - читает source, фильтрует НачислениеЗарплаты
    - группирует по (Подразделение, Сотрудник), суммирует Оклад, Премия, АдминистративныеРасходы
    - считает сотрудников с админ-расходами == threshold
    - записывает значение в ячейку
    """
    threshold = config.get("admin_threshold", 0)
    df = pd.read_excel(source_path, header=0)
    logger.info(f"Admin count: загружено {len(df)} строк")

    # Найти колонки
    def _find_col(df, keywords):
        for col in df.columns:
            c = col.lower().strip()
            if any(k in c for k in keywords):
                return col
        return None

    type_col = _find_col(df, ["типзаписи", "тип записи", "recordtype"])
    emp_col = _find_col(df, ["сотрудник", "employee", "фио"])
    dept_col = _find_col(df, ["подразделение", "department"])
    oklad_col = _find_col(df, ["оклад", "salary"])
    premia_col = _find_col(df, ["премия", "bonus"])
    admin_col = _find_col(df, ["административныерасходы", "administrative"])
    payroll_col = _find_col(df, ["видначислениязп", "payrolltype"])

    if not type_col:
        raise ValueError("Не найдена колонка ТипЗаписи")
    if not emp_col:
        raise ValueError("Не найдена колонка Сотрудник")
    if not admin_col:
        raise ValueError("Не найдена колонка АдминистративныеРасходы")

    # Фильтр: НачислениеЗарплаты
    df_payroll = df[df[type_col].astype(str).str.strip() == "НачислениеЗарплаты"]
    if df_payroll.empty:
        logger.warning("Нет строк с ТипЗаписи='НачислениеЗарплаты'")
        count = 0
    else:
        # Группировка: (Подразделение, Сотрудник), суммы
        group_keys = []
        if dept_col:
            group_keys.append(dept_col)
        if emp_col:
            group_keys.append(emp_col)

        if not group_keys:
            raise ValueError("Нет колонок для группировки")

        agg_funcs = {}
        if oklad_col and oklad_col in df_payroll.columns:
            agg_funcs[oklad_col] = "sum"
        if premia_col and premia_col in df_payroll.columns:
            agg_funcs[premia_col] = "sum"
        if admin_col and admin_col in df_payroll.columns:
            agg_funcs[admin_col] = "sum"

        df_grouped = df_payroll.groupby(group_keys, as_index=False).agg(agg_funcs)

        # Подсчёт: admin == threshold
        if threshold == 0:
            mask = (df_grouped[admin_col].isna()) | (df_grouped[admin_col] == 0)
        else:
            mask = df_grouped[admin_col] == threshold
        count = int(mask.sum())
        logger.info(
            f"Admin count threshold={threshold}: {count} сотрудников "
            f"(всего {len(df_grouped)} уникальных)"
        )

    # Открываем шаблон
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception:
        import tempfile
        import msoffcrypto
        with open(template_path, 'rb') as f:
            file = msoffcrypto.OfficeFile(f)
            if file.is_encrypted():
                file.load_key(password="987456")
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    file.decrypt(tmp)
                    tmp_path = tmp.name
                wb = openpyxl.load_workbook(tmp_path)
                os.unlink(tmp_path)
            else:
                raise

    ws = wb["Отчетность БИТ 2026"]

    # Целевая колонка (текущий месяц, offset=0)
    target_month = get_target_month_name(month, config.get("month_offset", 0))
    col_idx = find_month_column(ws, target_month, config.get("month_row", 2))
    if col_idx is None:
        wb.close()
        raise ValueError(
            f"Не найден месяц '{target_month}' в строке {config.get('month_row', 2)}"
        )

    # Целевая строка
    row_num = find_row_by_name(
        ws, config.get("row_column", 3), config.get("row_name", "")
    )
    if row_num is None:
        wb.close()
        raise ValueError(
            f"Не найдена строка '{config.get('row_name')}' "
            f"в колонке {get_column_letter(config.get('row_column', 3))}"
        )

    # Запись
    cell_ref = f"{get_column_letter(col_idx)}{row_num}"
    ws.cell(row=row_num, column=col_idx).value = count

    from src.excel_processor import _save_and_fix_formats
    await _save_and_fix_formats(wb, template_path)

    logger.info(f"Записано {config.get('key')}: {count} в {cell_ref}")
    return {
        "value": count,
        "cell": cell_ref,
        "label": config.get("label", ""),
        "format": config.get("format", "integer")
    }


def _calculate_ffot_sum(
    source_path: str,
    exclude_employee: Optional[dict] = None
) -> float:
    """
    Читает исходный файл, фильтрует ФактическийФОТ,
    суммирует колонку Сумма с исключениями.

    exclude_employee = {"Сорокин Илья Вячеславович": ["Премия"]}
    означает: для указанного сотрудника исключить строки с указанным видом начисления.
    """
    df = pd.read_excel(source_path, header=0)
    logger.info(f"FFOT: загружено {len(df)} строк, колонки: {list(df.columns)}")

    # Колонка типа записи
    type_col = None
    for col in df.columns:
        if col.lower().strip() in ("типзаписи", "тип записи", "тип", "recordtype"):
            type_col = col
            break
    if not type_col:
        raise ValueError("Не найдена колонка с типом записи (ТипЗаписи)")

    # Фильтр ФактическийФОТ
    df_ffot = df[df[type_col].astype(str).str.strip() == "ФактическийФОТ"]
    if df_ffot.empty:
        logger.warning("Нет строк с ТипЗаписи='ФактическийФОТ'")
        return 0.0

    # Колонки: Сумма, Сотрудник, ВидНачисления
    amount_col = emp_col = payroll_col = None
    for col in df_ffot.columns:
        c = col.lower().strip()
        if amount_col is None and c in ("сумма", "amount"):
            amount_col = col
        if emp_col is None and c in ("сотрудник", "employee", "фио"):
            emp_col = col
        if payroll_col is None and c in ("видначислениязп", "payrolltype"):
            payroll_col = col

    if not amount_col:
        raise ValueError("Не найдена колонка 'Сумма' в исходных данных")

    total = 0.0

    # Суммируем построчно с учётом исключений
    for _, row in df_ffot.iterrows():
        employee = str(row.get(emp_col, "")).strip() if emp_col else ""
        payroll_type = str(row.get(payroll_col, "")).strip() if payroll_col else ""

        # Проверяем исключения: для Сорокина не суммируем Премию
        skip_row = False
        if exclude_employee and employee in exclude_employee:
            excluded_types = exclude_employee[employee]
            if payroll_type in excluded_types:
                logger.info(
                    f"FFOT: исключена строка для {employee} ({payroll_type})"
                )
                skip_row = True

        if not skip_row:
            total += _clean_numeric(row.get(amount_col, 0))

    logger.info(f"FFOT: итог (по колонке Сумма) = {total}")
    return total


# ===================== READ VALUES =====================


def read_value_from_template(
    template_path: str,
    config: dict,
    month,
    year: int
) -> dict:
    """
    Читает значение из шаблона по конфигурации read_value.

    config = {
        "key": "profit",
        "label": "Прибыль отдела, мес, к зачету",
        "row_name": "Прибыль отдела, мес, к зачету",
        "row_column": 3,
        "month_row": 2,
        "month_offset": 0
    }
    """
    wb, ws = open_workbook_and_sheet(template_path, data_only=True)

    # Целевая колонка (текущий месяц при offset=0)
    target_month = get_target_month_name(month, config.get("month_offset", 0))
    col_idx = find_month_column(ws, target_month, config.get("month_row", 2))
    if col_idx is None:
        wb.close()
        raise ValueError(
            f"Не найден месяц '{target_month}' в строке {config.get('month_row', 2)}"
        )

    # Целевая строка
    row_num = find_row_by_name(
        ws, config.get("row_column", 3), config.get("row_name", "")
    )
    if row_num is None:
        wb.close()
        raise ValueError(
            f"Не найдена строка '{config.get('row_name')}' "
            f"в колонке {get_column_letter(config.get('row_column', 3))}"
        )

    # Чтение (с учётом объединённых ячеек)
    value = read_cell_at(ws, row_num, col_idx)
    wb.close()

    cell_ref = f"{get_column_letter(col_idx)}{row_num}"
    logger.info(f"Прочитано {config.get('key')}: {value} из {cell_ref}")

    return {
        "key": config.get("key"),
        "label": config.get("label", ""),
        "value": value,
        "cell": cell_ref,
        "format": config.get("format", "amount")
    }


async def process_write_values(
    source_path: str,
    template_path: str,
    write_values: List[dict],
    month,
    year: int
) -> list:
    """Обрабатывает все write_values из конфига."""
    results = []
    for cfg in write_values:
        try:
            if cfg.get("type") == "ffot":
                result = await _write_ffot(
                    source_path, template_path, cfg, month, year
                )
                results.append(result)
            elif cfg.get("type") == "admin_count":
                result = await _write_admin_count(
                    source_path, template_path, cfg, month, year
                )
                results.append(result)
            else:
                logger.warning(f"Неизвестный тип write_value: {cfg.get('type')}")
        except Exception as e:
            logger.exception(f"Ошибка write_value {cfg.get('key')}: {e}")
            results.append({"key": cfg.get("key"), "error": str(e)})
    return results


def process_read_values(
    template_path: str,
    read_values: List[dict],
    month,
    year: int
) -> list:
    """Обрабатывает все read_values из конфига."""
    results = []
    for cfg in read_values:
        try:
            result = read_value_from_template(
                template_path, cfg, month, year
            )
            results.append(result)
        except Exception as e:
            logger.exception(f"Ошибка read_value {cfg.get('key')}: {e}")
            results.append({
                "key": cfg.get("key"),
                "label": cfg.get("label", ""),
                "error": str(e)
            })
    return results
