"""
Установка автофильтра по колонке «Период» на листах-источниках.

После заполнения данных на листах ФОТ/ДС/Реализация/Взаиморасчеты
устанавливает фильтр по периоду, чтобы при открытии файла были
видны только строки текущего месяца.

Не использует LibreOffice UNO — только openpyxl.
Фильтр настраивается через auto_filter.ref + add_filter_column,
а строки других периодов физически скрываются через row_dimensions.
"""
import os
import logging

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Возможные названия колонки с периодом на разных листах
PERIOD_COLUMN_NAMES = ["Период", "ПериодЗаполнения"]


async def apply_period_filter(
    file_path: str,
    sheets: list[str],
    month: str,
    year: int
) -> dict:
    """
    Устанавливает автофильтр по колонке «Период» на указанных листах.

    1. Находит колонку «Период» по заголовку в строке 2
    2. Устанавливает auto_filter.ref на диапазон данных
    3. Добавляет критерий: показывать только строки с target_period
    4. Физически скрывает строки с НЕ target_period
    5. Сохраняет через _save_and_fix_formats() (LibreOffice конвертация)

    Аргументы:
        file_path: путь к Excel-файлу
        sheets: список имён листов для фильтрации
        month: название месяца (напр. «Май»)
        year: год (напр. 2026)

    Возвращает:
        {"status": "success", "result": {"filtered_sheets": [...], ...}}
        или {"status": "error", "error_message": "..."}
    """
    if not os.path.exists(file_path):
        return {"status": "error", "error_message": f"File not found: {file_path}"}

    target_period = f"{month} {year}"
    filtered_sheets: list[str] = []
    errors: list[str] = []

    # ======== Шаг 1: Устанавливаем фильтры ========
    wb = load_workbook(file_path)

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            errors.append(f"Sheet '{sheet_name}' not found")
            logger.warning(f"Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]

        # Ищем колонку «Период» во второй строке (заголовок)
        period_col_idx = _find_period_column(ws)
        if period_col_idx is None:
            errors.append(f"Sheet '{sheet_name}': column 'Период' not found in row 2")
            logger.warning(f"Sheet '{sheet_name}': period column not found, skipping")
            continue

        max_row = ws.max_row
        if max_row < 3:
            # Только заголовок, данных нет — фильтр не нужен
            logger.info(f"Sheet '{sheet_name}': no data rows, filter skipped")
            filtered_sheets.append(sheet_name)
            continue

        # Определяем последнюю колонку с заголовком
        actual_max_col = _find_last_header_column(ws) or ws.max_column

        # Диапазон фильтра: строка 2 (заголовок) → последняя строка данных
        filter_ref = f"A2:{get_column_letter(actual_max_col)}{max_row}"
        ws.auto_filter.ref = filter_ref

        # Критерий фильтра: показывать только строки target_period
        ws.auto_filter.add_filter_column(
            period_col_idx - 1,  # 0-индекс для openpyxl
            [target_period]
        )

        # Физически скрываем строки с другими периодами
        hidden_count = 0
        for row_idx in range(3, max_row + 1):
            cell_value = ws.cell(row=row_idx, column=period_col_idx).value
            if cell_value is not None and str(cell_value).strip() != target_period:
                ws.row_dimensions[row_idx].hidden = True
                hidden_count += 1

        logger.info(
            f"Sheet '{sheet_name}': filter='{target_period}' "
            f"range={filter_ref}, hidden={hidden_count} rows"
        )
        filtered_sheets.append(sheet_name)

    wb.close()

    if not filtered_sheets:
        return {
            "status": "error",
            "error_message": f"No sheets were filtered: {'; '.join(errors)}"
        }

    # ======== Шаг 2: Сохраняем напрямую через openpyxl ========
    # Не используем _save_and_fix_formats (LibreOffice), т.к. он сбрасывает auto_filter.
    # К этому моменту файл уже обработан LO предыдущими шагами (apply_sheet_mapping,
    # process_write_values), поэтому форматы в порядке.
    try:
        wb2 = load_workbook(file_path)
        wb2.save(file_path)
        wb2.close()
        logger.info(f"Filtered file saved: {file_path}")
    except Exception as e:
        logger.exception("Failed to save filtered file")
        return {
            "status": "error",
            "error_message": f"Save failed: {str(e)}"
        }

    result: dict = {
        "filtered_sheets": filtered_sheets,
        "target_period": target_period
    }
    if errors:
        result["warnings"] = errors

    logger.info(f"Period filter applied to sheets: {filtered_sheets}")
    return {"status": "success", "result": result}


def _find_period_column(ws) -> int | None:
    """Ищет колонку «Период» во второй строке листа. Возвращает 1-индекс или None."""
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col_idx).value
        if cell_value is not None and str(cell_value).strip() in PERIOD_COLUMN_NAMES:
            return col_idx
    return None


def _find_last_header_column(ws) -> int | None:
    """Ищет последнюю колонку с непустым заголовком во 2й строке."""
    last_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=2, column=col_idx).value is not None:
            last_col = col_idx
    return last_col
