import os
import shutil
import logging
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import msoffcrypto
import re

logger = logging.getLogger(__name__)

# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========

async def read_excel_structure(file_path: str) -> dict:
    try:
        if os.path.exists(file_path):
            try:
                with open(file_path, 'rb') as f:
                    file = msoffcrypto.OfficeFile(f)
                    if file.is_encrypted():
                        try:
                            file.load_key(password="987456")
                            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                                file.decrypt(tmp)
                                tmp_path = tmp.name
                            wb = load_workbook(tmp_path, data_only=True)
                            os.unlink(tmp_path)
                        except:
                            return {"status": "error", "error_message": "Cannot decrypt file with default password"}
                    else:
                        wb = load_workbook(file_path, data_only=True)
            except:
                wb = load_workbook(file_path, data_only=True)
        else:
            return {"status": "error", "error_message": "File not found"}
        
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = 2
            headers = [cell.value for cell in ws[header_row] if cell.value]
            sheets[sheet_name] = {
                "headers": headers,
                "rows_count": ws.max_row,
                "max_column": ws.max_column,
                "header_row": header_row
            }
        return {"status": "success", "result": {"sheets": sheets}}
    except Exception as e:
        logger.exception("read_excel_structure error")
        return {"status": "error", "error_message": str(e)}

def clean_number(value):
    if isinstance(value, str):
        cleaned = re.sub(r'[^\d,.-]', '', value.replace('\u00a0', '').replace(' ', ''))
        cleaned = cleaned.replace(',', '.')
        try:
            return float(cleaned)
        except ValueError:
            return value
    return value

# ========== ФУНКЦИЯ ДЛЯ ВЗАИМОРАСЧЕТОВ ==========

def _preprocess_vzaimoraschety(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """
    Предобработка данных для листа 'Взаиморасчеты':
    - вычисление направления (из поля direction_field, по умолчанию 'Направление') 
      Доход -> Исходящие, Расход -> Входящие
    - обработка спецофисов: строки с ключевым словом в ПодразделениеКонтрагентДляОтчета
      делятся по ТипОборота (БДР/БДДС) на "по актам" и "по ДС"
    - пустые ПодразделениеКонтрагентДляОтчета заполняются аналогично спецофисам,
      если задан ПодразделениеКонтрагент
    - схлопывание дублей БДР/БДДС для строк, НЕ содержащих ключевое слово
    """
    direction_field = config.get("direction_field", "Направление")
    turnover_field = config.get("turnover_type_field", "ТипОборота")
    
    dir_map = config.get("direction_mapping", {})
    # Создаём новую колонку с вычисленным направлением (перезаписываем direction_field)
    df["Направление"] = df[direction_field].map(dir_map)
    unmapped = df[df["Направление"].isna()][direction_field].unique()
    if len(unmapped) > 0:
        raise ValueError(f"Не удалось определить направление для значений {list(unmapped)} в поле '{direction_field}'")

    office_col = "ПодразделениеКонтрагентДляОтчета"
    keyword = config.get("special_office_keyword", "")
    replace_rules = config.get("special_office_replace", {})

    if turnover_field not in df.columns:
        raise ValueError(f"Для обработки Взаиморасчетов нужна колонка '{turnover_field}'")
    if "ПодразделениеКонтрагент" not in df.columns:
        raise ValueError("Для обработки Взаиморасчетов нужна колонка 'ПодразделениеКонтрагент'")

    # 1. Заполненные спецофисы (содержат ключевое слово) – разделяем по БДР/БДДС
    mask_special_filled = df[office_col].astype(str).str.contains(keyword, na=False)
    if mask_special_filled.any():
        df.loc[mask_special_filled, office_col] = df.loc[mask_special_filled, turnover_field].map(replace_rules)
        still_na = df.loc[mask_special_filled, office_col].isna()
        if still_na.any():
            bad_vals = df.loc[mask_special_filled & still_na, turnover_field].unique()
            raise ValueError(f"Для спецофиса не задана замена для {turnover_field}: {list(bad_vals)}")

    # 2. Пустые office_col – заполняем по тому же принципу (из ПодразделениеКонтрагент)
    mask_empty = df[office_col].isna() | (df[office_col].astype(str).str.strip() == "")
    if mask_empty.any():
        sub_office_col = "ПодразделениеКонтрагент"
        empty_without_sub = mask_empty & (df[sub_office_col].isna() | (df[sub_office_col].astype(str).str.strip() == ""))
        if empty_without_sub.any():
            problem_idx = df[empty_without_sub].index.tolist()
            raise ValueError(f"Пустой office_col и пустой ПодразделениеКонтрагент в строках: {problem_idx}")
        vid = df.loc[mask_empty, turnover_field]
        valid_vids = list(replace_rules.keys())
        invalid_vids = vid[~vid.isin(valid_vids)]
        if not invalid_vids.empty:
            raise ValueError(f"Некорректный {turnover_field} для пустых офисов: {invalid_vids.unique()}")
        df.loc[mask_empty, office_col] = vid.map(replace_rules)

    # 3. Схлопывание для обычных офисов (не содержащих keyword)
    is_special = df[office_col].astype(str).str.contains(keyword, na=False)
    df_special = df[is_special].copy()
    df_regular = df[~is_special].copy()

    if not df_regular.empty:
        group_keys = ["Подразделение", "Контрагент", "Проект", office_col, "Направление"]
        df_regular = df_regular.groupby(group_keys, as_index=False).first()
        # Удаляем служебные колонки, которые больше не нужны
        for col in [turnover_field, "ПодразделениеКонтрагент"]:
            if col in df_regular.columns:
                df_regular.drop(columns=[col], inplace=True)

    df = pd.concat([df_regular, df_special], ignore_index=True)

    # Убираем оставшиеся служебные колонки, но НЕ трогаем "Направление" (direction_field)
    cols_to_drop = [c for c in [turnover_field, "ПодразделениеКонтрагент"] if c in df.columns]
    # Также удаляем исходный direction_field только если он отличается от "Направление"
    if direction_field != "Направление" and direction_field in df.columns:
        cols_to_drop.append(direction_field)
    if cols_to_drop:
        df.drop(columns=cols_to_drop, inplace=True)

    return df

# ========== ОСНОВНАЯ ФУНКЦИЯ ==========

async def apply_sheet_mapping(source_path: str, template_path: str, sheet_name: str, mapping: dict, month: str, year: int, password: str = "987456", output_path: str = None) -> dict:
    try:
        if output_path is None:
            output_dir = os.path.dirname(template_path)
            output_filename = f"ДКП_10_-_{month}_{year}.xlsx"
            output_path = os.path.join(output_dir, output_filename)
            shutil.copy2(template_path, output_path)
            logger.info(f"Created new output file: {output_path}")
            
            if password:
                try:
                    with open(output_path, 'rb') as f:
                        file = msoffcrypto.OfficeFile(f)
                        if file.is_encrypted():
                            file.load_key(password=password)
                            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                                file.decrypt(tmp)
                                tmp_path = tmp.name
                            shutil.move(tmp_path, output_path)
                            logger.info("File decrypted")
                except Exception as e:
                    logger.warning(f"Decrypt failed: {e}")
        else:
            if not os.path.exists(output_path):
                return {"status": "error", "error_message": f"Output file not found: {output_path}"}
            logger.info(f"Using existing output file: {output_path}")

        # ======== ЧТЕНИЕ И ОЧИСТКА ========
        df_source = pd.read_excel(source_path, header=0)
        logger.info(f"Source rows before filters: {len(df_source)}")

        # Очистка денежных колонок: убираем неразрывные пробелы, запятые, превращаем в числа.
        # СтавкаНДС исключается, т.к. может содержать проценты или текст "Без НДС".
        money_keywords = ['Оклад', 'Премия', 'Сумма', 'Ставка', 'НДС']
        for col in df_source.columns:
            if col == "СтавкаНДС":
                continue
            if any(kw in col for kw in money_keywords):
                df_source[col] = (
                    df_source[col]
                    .astype(str)
                    .str.replace(r'[\s\u00a0]', '', regex=True)
                    .str.replace(',', '.')
                )
                df_source[col] = pd.to_numeric(df_source[col], errors='coerce')

        # ======== ФИЛЬТРЫ (включая списки) ========
        filters = mapping.get("filters", {})
        exclude_filters = mapping.get("exclude_filters", {})

        if exclude_filters:
            for col, val in exclude_filters.items():
                if col in df_source.columns:
                    if isinstance(val, list):
                        df_source = df_source[~df_source[col].astype(str).isin([str(v) for v in val])]
                        logger.info(f"Applied exclude filter {col} NOT IN {val}, rows left: {len(df_source)}")
                    else:
                        df_source = df_source[df_source[col].astype(str) != str(val)]
                        logger.info(f"Applied exclude filter {col} != {val}, rows left: {len(df_source)}")
                else:
                    logger.warning(f"Exclude column '{col}' not found in source")

        if filters:
            for col, val in filters.items():
                if col in df_source.columns:
                    if isinstance(val, list):
                        df_source = df_source[df_source[col].astype(str).isin([str(v) for v in val])]
                        logger.info(f"Applied filter {col} IN {val}, rows left: {len(df_source)}")
                    else:
                        df_source = df_source[df_source[col].astype(str) == str(val)]
                        logger.info(f"Applied filter {col} = {val}, rows left: {len(df_source)}")
                else:
                    logger.warning(f"Filter column '{col}' not found in source")

        if df_source.empty:
            return {"status": "error", "error_message": "No data after applying filters"}

        # ======== КАСТОМНАЯ ОБРАБОТКА (ВЗАИМОРАСЧЕТЫ) ========
        custom = mapping.get("custom_processing")
        if custom and custom.get("type") == "vzaimoraschety":
            df_source = _preprocess_vzaimoraschety(df_source, custom)

        if df_source.empty:
            return {"status": "error", "error_message": "No data after custom processing"}

        # ======== VIEW FILTERS ========
        view_filters = mapping.get("view_filters", {})
        if view_filters:
            for col, allowed_values in view_filters.items():
                if col in df_source.columns:
                    df_source = df_source[df_source[col].astype(str).isin([str(v) for v in allowed_values])]
                    logger.info(f"Applied view filter {col} in {allowed_values}, rows left: {len(df_source)}")
                else:
                    logger.warning(f"View filter column '{col}' not found in source")
            if df_source.empty:
                return {"status": "error", "error_message": "No data after applying view filters"}

        # ======== ГРУППИРОВКА И АГРЕГАЦИЯ ========
        group_by = mapping.get("group_by")
        aggregations = mapping.get("aggregations", {})

        if group_by:
            if isinstance(group_by, str):
                group_by = [group_by]
            group_by = [col for col in group_by if col in df_source.columns]
            if not group_by:
                data_for_insert = df_source.to_dict(orient='records')
            else:
                if "concat_premia" in aggregations.values():
                    def group_agg(group):
                        result = {}
                        for col in group_by:
                            result[col] = group[col].iloc[0]
                        if 'Оклад' in df_source.columns:
                            ok_labor = group.loc[group['ВидНачисленияЗП'] == 'Оплата труда', 'Оклад'].sum()
                            result['Оклад'] = ok_labor
                        else:
                            result['Оклад'] = 0
                        if 'Премия' in df_source.columns:
                            result['Премия'] = group['Премия'].sum()
                        else:
                            result['Премия'] = 0
                        if 'Комментарий' in df_source.columns:
                            comments = group.loc[group['ВидНачисленияЗП'] == 'Премия', 'Комментарий'].dropna().astype(str).unique()
                            result['Комментарий'] = "; ".join(comments)
                        else:
                            result['Комментарий'] = ""
                        return pd.Series(result)

                    df_grouped = df_source.groupby(group_by, as_index=False).apply(group_agg).reset_index(drop=True)
                    data_for_insert = df_grouped.to_dict(orient='records')
                else:
                    agg_funcs = {}
                    for agg_col, agg_type in aggregations.items():
                        if agg_col in df_source.columns:
                            if agg_type == "sum":
                                agg_funcs[agg_col] = "sum"
                            elif agg_type == "concat":
                                agg_funcs[agg_col] = lambda x: "; ".join(x.dropna().astype(str).unique())
                            else:
                                agg_funcs[agg_col] = agg_type
                    if agg_funcs:
                        df_grouped = df_source.groupby(group_by, as_index=False).agg(agg_funcs)
                        data_for_insert = df_grouped.to_dict(orient='records')
                    else:
                        data_for_insert = df_source.to_dict(orient='records')
        else:
            data_for_insert = df_source.to_dict(orient='records')

        # ======== ВСТАВКА В ШАБЛОН ========
        wb = load_workbook(output_path)
        if sheet_name not in wb.sheetnames:
            return {"status": "error", "error_message": f"Sheet '{sheet_name}' not found in template"}
        ws = wb[sheet_name]

        header_row = mapping.get("header_row", 2)
        source_cols = mapping.get("source_columns", {})
        source_cols = {int(k): v for k, v in source_cols.items()}
        append = mapping.get("append_to_end", True)

        rows_to_insert = []
        for row_dict in data_for_insert:
            new_row = {}
            for target_col_idx, source_expr in source_cols.items():
                if source_expr == "{month} {year}":
                    new_row[target_col_idx] = f"{month} {year}"
                else:
                    if source_expr in row_dict:
                        val = row_dict[source_expr]
                        # clean_number только для денежных колонок
                        if any(keyword in source_expr for keyword in ['Оклад', 'Премия', 'Сумма', 'Ставка', 'НДС']) and source_expr != "СтавкаНДС":
                            val = clean_number(val)
                        new_row[target_col_idx] = val
                    else:
                        new_row[target_col_idx] = None
            rows_to_insert.append(new_row)

        if not rows_to_insert:
            return {"status": "error", "error_message": "No data to insert"}

        logger.info(f"Prepared {len(rows_to_insert)} rows for insertion")

        if append:
            last_row = ws.max_row
            while last_row > 0:
                row_has_data = False
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=last_row, column=col).value is not None:
                        row_has_data = True
                        break
                if row_has_data:
                    break
                last_row -= 1
            start_row = last_row + 1
            if start_row <= header_row:
                start_row = header_row + 1
        else:
            if ws.max_row > header_row:
                ws.delete_rows(header_row + 1, ws.max_row - header_row)
            start_row = header_row + 1

        logger.info(f"Start row: {start_row}, number of rows to insert: {len(rows_to_insert)}")

        for i, row_dict in enumerate(rows_to_insert, start=start_row):
            for col_idx, value in row_dict.items():
                if value is not None:
                    ws.cell(row=i, column=col_idx).value = value

        wb.save(output_path)
        logger.info(f"File saved: {output_path}")

        return {"status": "success", "output_path": output_path, "rows_added": len(rows_to_insert)}
    except Exception as e:
        logger.exception("apply_sheet_mapping error")
        return {"status": "error", "error_message": str(e)}