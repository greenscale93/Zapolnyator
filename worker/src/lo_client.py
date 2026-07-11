"""
Клиент для работы с Excel: запись через openpyxl + фикс форматов через LibreOffice.

openpyxl используется ТОЛЬКО для записи данных (append rows, write cell).
После каждой записи файл конвертируется через LibreOffice --convert-to,
чтобы форматы дат и формулы не повреждались.

Чтение — через python-calamine (template_reader.py).

Почему не UNO: импорт `uno` в Docker с python:3.11-slim нестабилен
(требует точного совпадения версий LibreOffice и системного Python).
Данный гибридный подход — самый надёжный.
"""
import os
import shutil
import logging
import tempfile
import subprocess
import asyncio
from typing import List, Dict, Any, Optional

import openpyxl
import msoffcrypto

logger = logging.getLogger(__name__)


class LoClient:
    """Запись Excel через openpyxl + LibreOffice конвертация."""

    def __init__(self):
        self._lock = asyncio.Lock()
        logger.info("LoClient initialized (openpyxl + LibreOffice mode)")

    async def open_document(self, path: str, password: str = None):
        """Открывает Excel-файл через openpyxl."""
        file_path = path
        tmp = None
        if password:
            try:
                with open(path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    if office_file.is_encrypted():
                        office_file.load_key(password=password)
                        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                        office_file.decrypt(tmp)
                        tmp.close()
                        file_path = tmp.name
                        logger.info(f"Decrypted: {file_path}")
            except Exception as e:
                logger.warning(f"Decrypt failed: {e}")

        async with self._lock:
            wb = openpyxl.load_workbook(file_path)
            if tmp and file_path != path:
                try:
                    os.unlink(file_path)
                except Exception:
                    pass
            return wb

    async def get_sheet(self, wb, sheet_name: str):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        return wb[sheet_name]

    async def write_cell(self, ws, col: int, row: int, value):
        """Записывает значение в ячейку (col, row — 0-based)."""
        async with self._lock:
            cell = ws.cell(row=row + 1, column=col + 1)
            if value is None:
                return
            if isinstance(value, (int, float)):
                cell.value = float(value)
            else:
                cell.value = str(value)

    async def find_last_row(self, ws, max_column: int = 30) -> int:
        """Находит последнюю использованную строку (0-based)."""
        last = 0
        for row_idx in range(1, ws.max_row + 1):
            has_data = False
            for col_idx in range(1, max_column + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    has_data = True
                    break
            if has_data:
                last = row_idx - 1  # 0-based
            elif last > 0 and row_idx > last + 6:
                break
        return last

    async def append_rows(
        self,
        ws,
        data: List[Dict[int, Any]],
        start_row: Optional[int] = None,
        header_row: int = 1
    ):
        """Добавляет строки на лист (col в data — 0-based)."""
        async with self._lock:
            if start_row is None:
                start_row = await self.find_last_row(ws) + 1
            if start_row < header_row:
                start_row = header_row

            for i, row_data in enumerate(data):
                r = start_row + i + 1  # 1-based for openpyxl
                for col_idx, value in row_data.items():
                    c = col_idx + 1  # 1-based for openpyxl
                    if value is not None:
                        ws.cell(row=r, column=c).value = value
            logger.info(f"Appended {len(data)} rows starting at row {start_row + 1}")
            return start_row

    async def save_document(self, wb, path: str):
        """Сохраняет через LibreOffice --convert-to для фикса форматов."""
        tmp_dir = tempfile.mkdtemp()
        tmp_file = os.path.join(tmp_dir, "temp.xlsx")
        try:
            async with self._lock:
                wb.save(tmp_file)

            result = subprocess.run(
                [
                    "libreoffice", "--headless",
                    "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                    "--outdir", tmp_dir, tmp_file
                ],
                capture_output=True, text=True, timeout=30
            )
            if result.returncode != 0:
                raise RuntimeError(f"LibreOffice failed: {result.stderr[:200]}")

            converted = os.path.join(tmp_dir, "temp.xlsx")
            if not os.path.exists(converted):
                raise RuntimeError("LibreOffice did not produce output")
            shutil.move(converted, path)
            logger.info(f"Saved with format fix: {path}")
        finally:
            try:
                shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:
                pass

    async def close_document(self, wb):
        try:
            wb.close()
        except Exception:
            pass


lo_client = LoClient()

