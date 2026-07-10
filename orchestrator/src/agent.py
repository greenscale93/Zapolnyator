import logging
import os
import json
import asyncio
import httpx
from openpyxl import load_workbook
from src.session_manager import SessionManager
from src.memory import MemoryStore
from src.worker_client import WorkerClient
from src.mapping_store import get_mapping_dict, set_mapping, get_all_mappings, delete_mapping

logger = logging.getLogger(__name__)

class OrchestratorAgent:
    def __init__(self, session_manager: SessionManager, memory_store: MemoryStore, worker_client: WorkerClient):
        self.session_manager = session_manager
        self.memory_store = memory_store
        self.worker_client = worker_client
        self.bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
        self.rules = self._load_rules()
        logger.info("OrchestratorAgent initialized")

    def _load_rules(self) -> dict:
        try:
            with open("/app/rules.json", "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Cannot load rules: {e}")
            return {}

    async def _send_telegram_message(self, user_id: int, text: str, file_path: str = None, reply_markup: dict = None):
        if not self.bot_token:
            logger.warning("TELEGRAM_BOT_TOKEN not set")
            return
        url = f"https://api.telegram.org/bot{self.bot_token}/sendMessage"
        payload = {"chat_id": user_id, "text": text, "parse_mode": "HTML"}
        if reply_markup:
            payload["reply_markup"] = reply_markup
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                await client.post(url, json=payload)
                if file_path and os.path.exists(file_path):
                    with open(file_path, 'rb') as f:
                        files = {'document': (os.path.basename(file_path), f, 'text/plain')}
                        await client.post(
                            f"https://api.telegram.org/bot{self.bot_token}/sendDocument",
                            data={'chat_id': user_id},
                            files=files
                        )
        except Exception as e:
            logger.error(f"Telegram send error: {e}")

    def _read_template_offices(self, template_path: str) -> list:
        """Читает список подразделений из листа 'Отчетность БИТ 2026', колонка C, строки 13-16 и 34-35"""
        wb = load_workbook(template_path, data_only=True)
        if "Отчетность БИТ 2026" not in wb.sheetnames:
            raise ValueError("В шаблоне отсутствует лист 'Отчетность БИТ 2026'")
        ws = wb["Отчетность БИТ 2026"]
        offices = set()
        for row in range(13, 17):
            cell = ws.cell(row=row, column=3).value
            if cell and str(cell).strip():
                offices.add(str(cell).strip())
        for row in range(34, 36):
            cell = ws.cell(row=row, column=3).value
            if cell and str(cell).strip():
                offices.add(str(cell).strip())
        wb.close()
        return sorted(list(offices))

    async def run_agent_cycle(self, task_id: str, resume_from_question: bool = False):
        logger.info(f"Starting agent cycle for task {task_id}")
        state = await self.session_manager.get_session(task_id)
        if not state:
            logger.error(f"Task {task_id} not found")
            return

        user_id = state.get("user_id")
        files = state.get("files", {})
        data_file_path = files.get("data")
        excel_path = files.get("excel")
        if not data_file_path or not excel_path:
            await self._set_error(task_id, "Missing file paths")
            await self._send_telegram_message(user_id, "❌ Ошибка: отсутствуют пути к файлам.")
            return

        month = state.get("month", "Май")
        year = state.get("year", 2026)

        sheets_mapping = self.rules.get("sheets", {})
        output_path = state.get("output_path")  # если уже был создан

        # Обработка остальных листов, если не возобновляем после вопроса
        if not resume_from_question:
            for sheet_name, mapping in sheets_mapping.items():
                if sheet_name == "Взаиморасчеты":
                    continue  # обработаем позже
                await self._send_telegram_message(user_id, f"📝 Заполняю лист: {sheet_name}...")
                result = await self.worker_client.apply_sheet_mapping({
                    "source_path": data_file_path,
                    "template_path": excel_path if output_path is None else output_path,
                    "sheet_name": sheet_name,
                    "mapping": mapping,
                    "month": month,
                    "year": year,
                    "password": "987456",
                    "output_path": output_path
                })
                if result.get("status") == "error":
                    await self._set_error(task_id, result.get("error_message"))
                    await self._send_telegram_message(user_id, f"❌ Ошибка: {result.get('error_message')}")
                    return
                output_path = result["result"]["output_path"]
                rows_added = result["result"].get("rows_added", "неизвестно")
                await self._send_telegram_message(user_id, f"✅ Лист {sheet_name} заполнен (добавлено строк: {rows_added}).")
                # Сохраняем output_path в сессии
                await self.session_manager.update_session(task_id, {"output_path": output_path})

        # Теперь лист "Взаиморасчеты" с особым маппингом
        if "Взаиморасчеты" in sheets_mapping:
            vz_mapping = sheets_mapping["Взаиморасчеты"]
            await self._send_telegram_message(user_id, "📝 Заполняю лист: Взаиморасчеты...")

            # Получаем список пустых контрагентов
            try:
                empty_contractors = await self.worker_client.get_empty_vz_contractors(data_file_path)
            except Exception as e:
                await self._set_error(task_id, str(e))
                await self._send_telegram_message(user_id, f"❌ Ошибка получения данных: {str(e)}")
                return

            saved_mapping = get_mapping_dict()
            unknown = [c for c in empty_contractors if c not in saved_mapping]

            if unknown and not resume_from_question:
                # Читаем доступные офисы из шаблона
                try:
                    offices = self._read_template_offices(excel_path)
                except Exception as e:
                    await self._set_error(task_id, f"Ошибка чтения шаблона: {str(e)}")
                    return

                # Сохраняем контекст: список неизвестных и офисы
                await self.session_manager.update_session(task_id, {
                    "status": "waiting_question",
                    "question": {
                        "type": "vz_office_mapping",
                        "contractors": unknown,
                        "offices": offices,
                        "current_idx": 0
                    }
                })
                # Отправляем первый вопрос
                contractor = unknown[0]
                reply_markup = {
                    "inline_keyboard": self._build_office_keyboard(offices, contractor)
                }
                await self._send_telegram_message(user_id, f"Выберите подразделение для контрагента «{contractor}»:",
                                                  reply_markup=reply_markup)
                return  # ждём ответа

            # Если нет неизвестных или возобновляем после всех ответов – выполняем apply_sheet_mapping
            # Передаём маппинг и custom_processing
            vz_mapping["custom_processing"]["office_mapping"] = saved_mapping  # добавим в mapping для worker
            result = await self.worker_client.apply_sheet_mapping({
                "source_path": data_file_path,
                "template_path": excel_path if output_path is None else output_path,
                "sheet_name": "Взаиморасчеты",
                "mapping": vz_mapping,
                "month": month,
                "year": year,
                "password": "987456",
                "output_path": output_path
            })
            if result.get("status") == "error":
                await self._set_error(task_id, result.get("error_message"))
                await self._send_telegram_message(user_id, f"❌ Ошибка: {result.get('error_message')}")
                return
            output_path = result["result"]["output_path"]
            rows_added = result["result"].get("rows_added", "неизвестно")
            await self._send_telegram_message(user_id, f"✅ Лист Взаиморасчеты заполнен (добавлено строк: {rows_added}).")
            await self.session_manager.update_session(task_id, {"output_path": output_path})

        # Завершение
        if output_path:
            await self._send_telegram_message(user_id, "✅ Обработка завершена! Файл будет отправлен.")
            await self.session_manager.update_session(task_id, {
                "status": "done",
                "result_file": output_path,
                "metrics": {}
            })
        else:
            await self._set_error(task_id, "No output file generated")

    def _build_office_keyboard(self, offices: list, contractor: str):
        """Создаёт inline-клавиатуру с офисами и кнопкой 'Чужой офис'"""
        keyboard = []
        # группируем по 2 в ряд
        row = []
        for office in offices:
            row.append({"text": office, "callback_data": f"map_office|{contractor}|{office}"})
            if len(row) == 2:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
        # кнопка "Чужой офис"
        keyboard.append([{"text": "Чужой офис", "callback_data": f"map_office|{contractor}|Взаиморасчеты с др. офисами"}])
        return keyboard

    async def handle_answer(self, task_id: str, answer_data: str):
        """Обработка inline-ответа на вопрос маппинга"""
        # answer_data формат: "map_office|contractor|selected_office"
        parts = answer_data.split("|")
        if len(parts) != 3 or parts[0] != "map_office":
            return
        contractor = parts[1]
        selected_office = parts[2]

        # Сохраняем маппинг
        set_mapping(contractor, selected_office)

        # Получаем состояние задачи
        state = await self.session_manager.get_session(task_id)
        question = state.get("question")
        if not question or question.get("type") != "vz_office_mapping":
            return

        contractors = question["contractors"]
        current_idx = question["current_idx"]
        # Убедимся, что текущий contractor соответствует индексу
        if current_idx < len(contractors) and contractors[current_idx] == contractor:
            current_idx += 1
            if current_idx < len(contractors):
                # Следующий вопрос
                next_contractor = contractors[current_idx]
                offices = question["offices"]
                reply_markup = {
                    "inline_keyboard": self._build_office_keyboard(offices, next_contractor)
                }
                user_id = state["user_id"]
                await self._send_telegram_message(user_id, f"Выберите подразделение для контрагента «{next_contractor}»:",
                                                  reply_markup=reply_markup)
                await self.session_manager.update_session(task_id, {
                    "status": "waiting_question",
                    "question": {
                        "type": "vz_office_mapping",
                        "contractors": contractors,
                        "offices": offices,
                        "current_idx": current_idx
                    }
                })
            else:
                # Больше вопросов нет – продолжаем обработку
                await self.session_manager.update_session(task_id, {"status": "running", "question": None})
                asyncio.create_task(self.run_agent_cycle(task_id, resume_from_question=True))
        else:
            logger.error(f"Answer mismatch: expected {contractors[current_idx]}, got {contractor}")

    async def edit_mapping_command(self, user_id: int):
        """Отправляет список текущих маппингов с возможностью удалить"""
        mappings = get_all_mappings()
        if not mappings:
            await self._send_telegram_message(user_id, "Нет сохранённых маппингов.")
            return

        keyboard = []
        for cont, rep in mappings:
            keyboard.append([{"text": f"{cont} → {rep} ❌", "callback_data": f"del_mapping|{cont}"}])
        keyboard.append([{"text": "Отмена", "callback_data": "edit_mapping_cancel"}])
        reply_markup = {"inline_keyboard": keyboard}
        await self._send_telegram_message(user_id, "Текущие маппинги. Нажмите для удаления:", reply_markup=reply_markup)

    async def handle_delete_mapping(self, task_id: str, contractor: str):
        delete_mapping(contractor)
        state = await self.session_manager.get_session(task_id)
        if state:
            await self._send_telegram_message(state["user_id"], f"Маппинг для {contractor} удалён.")
        # обновить список? можно отправить заново /edit_mapping

    async def stop_task(self, task_id: str):
        await self.session_manager.update_session(task_id, {
            "status": "cancelled",
            "error": "Остановлено пользователем"
        })
        logger.info(f"Task {task_id} stopped by user")

    async def _set_error(self, task_id: str, error_message: str):
        logger.error(f"Task {task_id} error: {error_message}")
        await self.session_manager.update_session(task_id, {
            "status": "error",
            "error": error_message
        })